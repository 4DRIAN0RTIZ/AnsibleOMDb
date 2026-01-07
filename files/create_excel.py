#!/usr/bin/env python3
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_excel(data, output_file):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Movie Results"

        # Header styling
        header_fill = PatternFill(start_color="e38528", end_color="e38528", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = ["Title", "Year", "imdbID", "Type", "Poster"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        movies = data.get("Search", [])
        for row_num, movie in enumerate(movies, 2):
            ws.cell(row=row_num, column=1, value=movie.get("Title", "N/A"))
            ws.cell(row=row_num, column=2, value=movie.get("Year", "N/A"))
            ws.cell(row=row_num, column=3, value=movie.get("imdbID", "N/A"))
            ws.cell(row=row_num, column=4, value=movie.get("Type", "N/A"))
            ws.cell(row=row_num, column=5, value=movie.get("Poster", "N/A"))

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 50

        wb.save(output_file)
        return True, "Excel file created successfully"
    except Exception as e:
        return False, f"Error creating Excel file: {str(e)}"

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(json.dumps({"success": False, "message": "Invalid arguments"}))
        sys.exit(1)

    data_file = sys.argv[1]
    output_file = sys.argv[2]

    try:
        with open(data_file, 'r') as f:
            data = json.load(f)

        success, message = create_excel(data, output_file)
        result = {"success": success, "message": message}
        print(json.dumps(result))
        sys.exit(0 if success else 1)
    except Exception as e:
        print(json.dumps({"success": False, "message": str(e)}))
        sys.exit(1)
